import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# 文件路径
input_path = 'input.xlsx'
output_path = 'output.xlsx'

# 1. 用 pandas 读取和处理数据
df = pd.read_excel(input_path)

# 拆分字段
df[['组织1', '组织2']] = df['组织'].str.split('/', expand=True)
df['组织1'] = df['组织1'].str.strip()
df['组织2'] = df['组织2'].str.strip()

event_type_split = df['事件类型'].str.split('/', expand=True)
for i in range(3):
    df[f'事件类型{i+1}'] = event_type_split[i].str.strip()

# 计算 Lead Time
df['报单时间'] = pd.to_datetime(df['报单时间'])
df['完成时间'] = pd.to_datetime(df['完成时间'])
df['Lead Time'] = (df['完成时间'] - df['报单时间']).dt.total_seconds() // 60

# 按顺序保留需要的字段
desired_columns = [
    '任务ID', '标题', '创建者', '执行者', '紧急程度', '影响级', '优先级',
    '事件来源', '联系人', '单量', 'Jira工单',
    '组织1', '组织2',
    '事件类型1', '事件类型2', '事件类型3',
    '报单时间', '完成时间', '是否完成', '借助伙伴资源',
    'Lead Time'
]

df_final = df[desired_columns]

# 保存为 Excel（表头默认已加粗）
df_final.to_excel(output_path, index=False)

# 2. 用 openpyxl 再处理格式
wb = load_workbook(output_path)
ws = wb.active

# 显式设置表头加粗（虽然 pandas 默认已经加了，但保险起见）
for cell in ws[1]:
    cell.font = Font(bold=True)

# ✅ 加 Filter（筛选器）
ws.auto_filter.ref = ws.dimensions

wb.save(output_path)

print("✅ 已输出文件，表头加粗并添加筛选器：", output_path)
